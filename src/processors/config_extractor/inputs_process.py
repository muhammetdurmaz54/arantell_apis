from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import io
import boto3
from botocore.exceptions import ClientError
import json
from tempfile import NamedTemporaryFile
import openpyxl
import zipfile
import pandas as pd
from src.processors.config_extractor.configurator import Configurator
'''
    ship_name - ship_imo/Logs or Noon/Engine or Fuel/ship_imo:logsornoon:engineorfuel.txt
    ATM Book - 9591301/Logs/Engine/9591301logsengine.txt
'''
class Inputs():
    def __init__(self, ship_imo, ship_name, type_of_input, subtype_of_input, aws_id, aws_secret, bucket_name):
        self.ship_imo = ship_imo
        self.ship_name = ship_name
        self.type_of_input = type_of_input
        self.subtype_of_input = subtype_of_input
        self.aws_id = aws_id
        self.aws_secret = aws_secret
        self.bucket_name = bucket_name
        self.file_name = str(ship_imo) + subtype_of_input + type_of_input + '.txt'
        self.object_key = ship_name + ' - ' + str(ship_imo) + '/' + subtype_of_input.capitalize() + '/' + type_of_input.capitalize() + '/' + self.file_name
    
    def getSpreadsheetFromS3(self, how_many_days='7'):
        configuration = Configurator(self.ship_imo)
        ship_configs = configuration.get_ship_configs()

        s3 = boto3.client('s3', aws_access_key_id=self.aws_id, aws_secret_access_key=self.aws_secret)
        # response = s3.list_objects(Bucket=bucket_name)
        # print(response)
        try:
            obj = s3.get_object(Bucket=self.bucket_name, Key=self.object_key)
        except:
            new_object_key = self.object_key.replace('txt', 'xlsx')
            # try:
            obj = s3.get_object(Bucket=self.bucket_name, Key=new_object_key)
            # except:
            #     return "Cannot Find Template!", 400
        data = obj['Body'].read()
        print(type(data))
        # new_data = data.decode('utf-8')
        # json_df_dict = json.dumps(new_data)
        try:
            json_df_dict = json.loads(data)

            sheet_names_list=[json_df_dict['sheets'][sheet]['name'] for sheet in range(len(json_df_dict['sheets']))]

            spreadsheet_metadata={}
            for sheet in range(len(sheet_names_list)):
                new_list=[]
                rowTemp={}
                if 'rows' in json_df_dict['sheets'][sheet]:
                    for i in range(len(json_df_dict['sheets'][sheet]['rows'])):
                        if i < 4:
                            print("ROW NUMBER!!!!!", i)
                            temp={}
                            for j in range(len(json_df_dict['sheets'][sheet]['rows'][i]['cells'])):
                                if json_df_dict['sheets'][sheet]['rows'][i]['cells'][j] != None and 'value' in json_df_dict['sheets'][sheet]['rows'][i]['cells'][j]:
                                    tempList=[]
                                    colSpan = json_df_dict['sheets'][sheet]['rows'][i]['cells'][j]['colSpan'] if 'colSpan' in json_df_dict['sheets'][sheet]['rows'][i]['cells'][j] else ""
                                    rowSpan = json_df_dict['sheets'][sheet]['rows'][i]['cells'][j]['rowSpan'] if 'rowSpan' in json_df_dict['sheets'][sheet]['rows'][i]['cells'][j] else ""
                                    styles = json_df_dict['sheets'][sheet]['rows'][i]['cells'][j]['style'] if 'style' in json_df_dict['sheets'][sheet]['rows'][i]['cells'][j] else {}
                                    tempList.append(colSpan)
                                    tempList.append(rowSpan)
                                    tempList.append(styles)
                                    temp[j] = tempList
                                    rowTemp[i] = temp
                                else:
                                    tempList=[]
                                    colSpan = ""
                                    rowSpan = ""
                                    styles = {}
                                    tempList.append(colSpan)
                                    tempList.append(rowSpan)
                                    tempList.append(styles)
                                    temp[j] = tempList
                            # new_list.append({i: temp})
                                    rowTemp[i] = temp
                                # print("ROW TEMP!!!!!!!!", rowTemp)
                            # if len(list(rowTemp.keys())) > 0:
                            spreadsheet_metadata[sheet_names_list[sheet]] = rowTemp
                            # else:
                            #     continue

            if self.subtype_of_input != 'logs':
                headers_dict={}
                for sheet in range(len(sheet_names_list)):
                    headers=[]
                    for i in range(len(json_df_dict['sheets'][sheet]['rows'][0]['cells'])):
                        # if i == 0:
                        if 'value' in json_df_dict['sheets'][sheet]['rows'][0]['cells'][i]:
                            headers.append(json_df_dict['sheets'][sheet]['rows'][0]['cells'][i]['value'])
                        else:
                            headers.append(None)
                    headers_dict[sheet_names_list[sheet]] = headers
                        # continue
                # length_of_json_df_dict = len(json_df_dict['sheets'][0]['rows'])
                # parameter_list_dict = json_df_dict['sheets'][0]['rows'][length_of_json_df_dict-1]['cells']
                # values=[]
                # for i in parameter_list_dict:
                #     if 'value' in i:
                #         values.append(i['value'])
                #     else:
                #         values.append(None)
                raw_dictionary = configuration.get_dict_of_raw_values(how_many_days, headers_dict)
                print("RAW DICTIONARY!!!!!", len(raw_dictionary))
                result={}
                for sheet in range(len(sheet_names_list)):
                    temp_list=[]
                    for i in range(0, len(json_df_dict['sheets'][sheet]['rows'])):
                        if i < 4:
                            temp_list.append(json_df_dict['sheets'][sheet]['rows'][i]['cells'])
                    result[sheet_names_list[sheet]] = temp_list
                print("LENGTH BEFORE", len(result))
                for sheet in raw_dictionary.keys():
                    for i in range(len(raw_dictionary[sheet])-1, -1, -1):
                        # print(raw_dictionary[i])
                        result[sheet].append(raw_dictionary[sheet][i])
                print("LENGTH AFTER", len(result))
                return result, spreadsheet_metadata, sheet_names_list
            else:
                # result=[]
                # for i in range(len(json_df_dict['sheets'][0]['rows'])):
                #     result.append(json_df_dict['sheets'][0]['rows'][i]['cells'])
                # return result, spreadsheet_metadata, sheet_names_list
                headers_dict={}
                for sheet in range(len(sheet_names_list)):
                    headers=[]
                    if 'rows' in json_df_dict['sheets'][sheet]:
                        for i in range(len(json_df_dict['sheets'][sheet]['rows'][0]['cells'])):
                            # if i == 0:
                            if 'value' in json_df_dict['sheets'][sheet]['rows'][0]['cells'][i]:
                                headers.append(json_df_dict['sheets'][sheet]['rows'][0]['cells'][i]['value'])
                            else:
                                headers.append(None)
                        headers_dict[sheet_names_list[sheet]] = headers
                        # continue
                # length_of_json_df_dict = len(json_df_dict['sheets'][0]['rows'])
                # parameter_list_dict = json_df_dict['sheets'][0]['rows'][length_of_json_df_dict-1]['cells']
                # values=[]
                # for i in parameter_list_dict:
                #     if 'value' in i:
                #         values.append(i['value'])
                #     else:
                #         values.append(None)
                raw_dictionary = configuration.get_dict_of_raw_values(how_many_days, headers_dict)
                print("RAW DICTIONARY!!!!!", len(raw_dictionary))
                result={}
                for sheet in range(len(sheet_names_list)):
                    temp_list=[]
                    if 'rows' in json_df_dict['sheets'][sheet]:
                        for i in range(0, len(json_df_dict['sheets'][sheet]['rows'])):
                            if i < 4:
                                temp_list.append(json_df_dict['sheets'][sheet]['rows'][i]['cells'])
                        result[sheet_names_list[sheet]] = temp_list
                print("LENGTH BEFORE", len(result))
                for sheet in raw_dictionary.keys():
                    for i in range(len(raw_dictionary[sheet])-1, -1, -1):
                        # print(raw_dictionary[i])
                        result[sheet].append(raw_dictionary[sheet][i])
                print("LENGTH AFTER", len(result))
                return result, spreadsheet_metadata, sheet_names_list
        except UnicodeDecodeError:
            new_excel = pd.ExcelFile(io.BytesIO(data), engine="openpyxl")
            # new_data = pd.read_excel(io.BytesIO(data), engine="openpyxl")
            new_data = pd.read_excel(new_excel, None)
            # json_df_dict = json.loads(new_data)
            # print(type(new_data))
            # print(new_data)
            for sheet in new_data: #{sheet_name: Dataframe}
                for column in new_data[sheet]:
                    if 'Unnamed' in column:
                        new_data[sheet] = new_data[sheet].drop(columns=column)
                new_data[sheet] = json.loads(new_data[sheet].to_json())
            # print(type(new_data))
            # print(new_data)

            print(new_data)
            sheet_names_list = list(new_data.keys())
            
            spreadsheet_metadata={}
            # for sheet in range(len(json_df_dict['sheets'])):
            for sheet in new_data:
                new_list=[]
                rowTemp={}
                for i in range(0,4):
                    temp={}
                    for j in range(len(list(new_data[sheet].keys()))):
                        # if i < 3:
                        print("ROW NUMBER!!!!!", i)
                        # for j in range(len(json_df_dict['sheets'][sheet]['rows'][i]['cells'])):
                            # if json_df_dict['sheets'][sheet]['rows'][i]['cells'][j] != None and 'value' in json_df_dict['sheets'][sheet]['rows'][i]['cells'][j]:
                            #     tempList=[]
                            #     colSpan = json_df_dict['sheets'][sheet]['rows'][i]['cells'][j]['colSpan'] if 'colSpan' in json_df_dict['sheets'][sheet]['rows'][i]['cells'][j] else ""
                            #     rowSpan = json_df_dict['sheets'][sheet]['rows'][i]['cells'][j]['rowSpan'] if 'rowSpan' in json_df_dict['sheets'][sheet]['rows'][i]['cells'][j] else ""
                            #     styles = json_df_dict['sheets'][sheet]['rows'][i]['cells'][j]['style'] if 'style' in json_df_dict['sheets'][sheet]['rows'][i]['cells'][j] else {}
                            #     tempList.append(colSpan)
                            #     tempList.append(rowSpan)
                            #     tempList.append(styles)
                            #     temp[j] = tempList
                            #     rowTemp[i] = temp
                            # else:
                        tempList=[]
                        colSpan = ""
                        rowSpan = ""
                        styles = {}
                        tempList.append(colSpan)
                        tempList.append(rowSpan)
                        tempList.append(styles)
                        temp[j] = tempList
                        # new_list.append({i: temp})
                        rowTemp[i] = temp
                            # print("ROW TEMP!!!!!!!!", rowTemp)
                        # if len(list(rowTemp.keys())) > 0:
                    spreadsheet_metadata[sheet] = rowTemp
                        # else:
                        #     continue

            if self.subtype_of_input != 'logs':
                headers_dict={}
                for sheet in new_data:
                    headers=[]
                    for header in new_data[sheet].keys():
                        headers.append(header)
                    headers_dict[sheet] = headers
                # for sheet in range(len(json_df_dict['sheets'])):
                #     headers=[]
                #     for i in range(len(json_df_dict['sheets'][sheet]['rows'][0]['cells'])):
                #         # if i == 0:
                #         if 'value' in json_df_dict['sheets'][sheet]['rows'][0]['cells'][i]:
                #             headers.append(json_df_dict['sheets'][sheet]['rows'][0]['cells'][i]['value'])
                #         else:
                #             headers.append(None)
                    # headers_dict[sheet] = headers
                        # continue
                # length_of_json_df_dict = len(json_df_dict['sheets'][0]['rows'])
                # parameter_list_dict = json_df_dict['sheets'][0]['rows'][length_of_json_df_dict-1]['cells']
                # values=[]
                # for i in parameter_list_dict:
                #     if 'value' in i:
                #         values.append(i['value'])
                #     else:
                #         values.append(None)

                raw_dictionary = configuration.get_dict_of_raw_values(how_many_days, headers_dict)
                print("RAW DICTIONARY!!!!!", len(raw_dictionary))
                result={}
                rowList = []
                for sheet in new_data:
                    temp_list=[]
                    for header in new_data[sheet]:
                        temp_list.append({'value': header})
                    rowList.append(temp_list)
                    result[sheet] = rowList
                    temp_list=[]
                    for header in new_data[sheet]:
                        temp_list.append({'value': new_data[sheet][header]['0']})
                    rowList.append(temp_list)
                    result[sheet] = rowList
                    temp_list=[]
                    for header in new_data[sheet]:
                        temp_list.append({'value': new_data[sheet][header]['1']})
                    rowList.append(temp_list)
                    result[sheet] = rowList
                    rowList=[]
                    for header in new_data[sheet]:
                        temp_list.append({'value': new_data[sheet][header]['2']})
                    rowList.append(temp_list)
                    result[sheet] = rowList
                    rowList=[]
                # for sheet in range(len(json_df_dict['sheets'])):
                #     temp_list=[]
                #     for i in range(0, len(json_df_dict['sheets'][sheet]['rows'])):
                #         if i < 3:
                #             temp_list.append(json_df_dict['sheets'][sheet]['rows'][i]['cells'])
                #     result[sheet] = temp_list
                print("LENGTH BEFORE", len(result))
                for sheet in raw_dictionary.keys():
                    for i in range(len(raw_dictionary[sheet])-1, -1, -1):
                        # print(raw_dictionary[i])
                        result[sheet].append(raw_dictionary[sheet][i])
                print("LENGTH AFTER", len(result))
                return result, spreadsheet_metadata, sheet_names_list
            else:
                # result=[]
                # for i in range(len(json_df_dict['sheets'][0]['rows'])):
                #     result.append(json_df_dict['sheets'][0]['rows'][i]['cells'])
                # return result, spreadsheet_metadata, sheet_names_list
                headers_dict={}
                for sheet in new_data:
                    headers=[]
                    for header in new_data[sheet].keys():
                        headers.append(header)
                    headers_dict[sheet] = headers
                # for sheet in range(len(json_df_dict['sheets'])):
                #     headers=[]
                #     for i in range(len(json_df_dict['sheets'][sheet]['rows'][0]['cells'])):
                #         # if i == 0:
                #         if 'value' in json_df_dict['sheets'][sheet]['rows'][0]['cells'][i]:
                #             headers.append(json_df_dict['sheets'][sheet]['rows'][0]['cells'][i]['value'])
                #         else:
                #             headers.append(None)
                    # headers_dict[sheet] = headers
                        # continue
                # length_of_json_df_dict = len(json_df_dict['sheets'][0]['rows'])
                # parameter_list_dict = json_df_dict['sheets'][0]['rows'][length_of_json_df_dict-1]['cells']
                # values=[]
                # for i in parameter_list_dict:
                #     if 'value' in i:
                #         values.append(i['value'])
                #     else:
                #         values.append(None)
                print("HEADER DICT!!!!!!", headers_dict)
                raw_dictionary = configuration.get_dict_of_raw_values(how_many_days, headers_dict)
                print("RAW DICTIONARY!!!!!", len(raw_dictionary))
                result={}
                rowList = []
                for sheet in new_data:
                    temp_list=[]
                    for header in new_data[sheet]:
                        temp_list.append({'value': header})
                    rowList.append(temp_list)
                    result[sheet] = rowList
                    temp_list=[]
                    for header in new_data[sheet]:
                        temp_list.append({'value': new_data[sheet][header]['0']})
                    rowList.append(temp_list)
                    result[sheet] = rowList
                    temp_list=[]
                    for header in new_data[sheet]:
                        temp_list.append({'value': new_data[sheet][header]['1']})
                    rowList.append(temp_list)
                    result[sheet] = rowList
                    rowList=[]
                # for sheet in range(len(json_df_dict['sheets'])):
                #     temp_list=[]
                #     for i in range(0, len(json_df_dict['sheets'][sheet]['rows'])):
                #         if i < 3:
                #             temp_list.append(json_df_dict['sheets'][sheet]['rows'][i]['cells'])
                #     result[sheet] = temp_list
                print("LENGTH BEFORE", len(result))
                for sheet in raw_dictionary.keys():
                    for i in range(len(raw_dictionary[sheet])-1, -1, -1):
                        # print(raw_dictionary[i])
                        result[sheet].append(raw_dictionary[sheet][i])
                print("LENGTH AFTER", len(result))
                return result, spreadsheet_metadata, sheet_names_list
    
    def putSpreadsheetToS3(self, body_json):
        # configuration = Configurator(self.ship_imo)
        # ship_configs = configuration.get_ship_configs()
        # print(body_json)
        # body_json_json = json.loads(body_json)
        # print(body_json_json)
        # iobody_json = io.BytesIO(body_json)
        # df = pd.read_excel(iobody_json)
        # print(iobody_json)
        # zipbody_json = zipfile.ZipFile(iobody_json.read())
        # with zipfile.ZipFile(iobody_json) as archive:
        # wb = load_workbook(body_json)
        # keys = list(body_json.keys())
        # wb = Workbook()
        # ws = wb.active
        # print(type(body_json))
        # # with open('users.json',encoding="utf8") as f:
        # #     json_data = json.load(f)
        
        # for i in range(len(keys)):
        #     ws.cell(row=1, column=i+1, value=keys[i])
        
        # for i in range(len(keys)):
        #     for j in range(len(body_json[keys[i]])):
        #         ws.cell(row=j+2, column=i+1, value=body_json[keys[i]][j])

        s3 = boto3.client('s3', aws_access_key_id=self.aws_id, aws_secret_access_key=self.aws_secret)
        # with open(body_json, 'rb') as f:
        try:
            s3.put_object(Body=body_json, Bucket=self.bucket_name, Key=self.object_key)
            return "Saved as Template! Refresh the page to see the changes."
        except ClientError as e:
            return {"Status": e.response}
        # try:
        #     with NamedTemporaryFile() as tmp:
        #         filename = self.file_name
        #         wb.save(filename)
        #     #     json_data = json.load(f)     bytearray(str(body_json), encoding="utf8")
        #         s3.upload_file(filename, self.bucket_name, self.object_key)
        # except ClientError as e:
        #     print(e)
        #     return False